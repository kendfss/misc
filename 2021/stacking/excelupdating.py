"""
https://stackoverflow.com/questions/66246080/writing-to-multiple-cells-in-excel-using-python-but-get-error-message-tuple-h
"""

# The cell_obj is referring to only one cell value and not to all 3 cells. You can set values by setting each cell.

import openpyxl, os
# wb_obj = openpyxl.load_workbook(latest_file)
wb_obj = openpyxl.load_workbook('book1.xlsx')
sheet_obj = wb_obj.active
max_column = sheet_obj.max_column

#for i in range(1, max_column+1):
#   cell_obj = sheet_obj.cell(row=1, column = i)
#   print(cell_obj.value)

# Save cell values into a list
my_list = [sheet_obj.cell(row=1, column = i).value for i in range(1, max_column+1)]
print(my_list)

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Sheet 1'
# my_list = cell_obj.value

# Iterate over list and update each cell
for i, v in enumerate(my_list):
    sheet.cell(row=i+1, column=1).value = v


name = 'hello_world.xlsx'
wb.save(name)
os.startfile(name)

